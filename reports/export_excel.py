"""
Xuất báo cáo Excel
"""

import os
from datetime import datetime
from typing import List, Dict

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("[WARN] openpyxl chưa được cài. Chạy: pip install openpyxl")


def _header_style(ws, row, col, value, bg_color="1E3A5F", font_color="FFFFFF"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=font_color, size=11)
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell


def _data_style(ws, row, col, value, bg_color=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="EEEEEE")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if bg_color:
        cell.fill = PatternFill("solid", fgColor=bg_color)
    return cell


def export_results(
    results: List[Dict],
    freq_data: Dict[str, int],
    gan_data: List[Dict],
    head_tail: Dict,
    output_path: str = None,
    start_date: str = "",
    end_date: str = "",
    province: str = "",
) -> str:
    """
    Xuất báo cáo Excel đầy đủ
    Returns: đường dẫn file đã tạo
    """
    if not OPENPYXL_OK:
        raise RuntimeError("Cần cài openpyxl: pip install openpyxl")

    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(
            os.path.expanduser("~"), "Desktop", f"BaoCaoXoSo_{ts}.xlsx"
        )

    wb = openpyxl.Workbook()

    # ── Sheet 1: Kết quả ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Kết Quả"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 12
    for col in "DEFGHIJK":
        ws1.column_dimensions[col].width = 22

    # Tiêu đề
    ws1.merge_cells("A1:K1")
    title_cell = ws1["A1"]
    title_cell.value = f"BÁO CÁO KẾT QUẢ — {province} — {start_date} đến {end_date}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1E3A5F")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    headers = ["Ngày", "Tỉnh", "ĐB", "Giải 1", "Giải 2", "Giải 3", "Giải 4", "Giải 5", "Giải 6", "Giải 7", "Giải 8"]
    for i, h in enumerate(headers, 1):
        _header_style(ws1, 2, i, h)
    ws1.row_dimensions[2].height = 22

    alt_colors = ["F8FAFE", "FFFFFF"]
    for r_idx, row in enumerate(results, 3):
        bg = alt_colors[r_idx % 2]
        _data_style(ws1, r_idx, 1, row.get("draw_date", ""), bg)
        _data_style(ws1, r_idx, 2, row.get("province", ""), bg)
        _data_style(ws1, r_idx, 3, row.get("special_prize", ""), bg)
        _data_style(ws1, r_idx, 4, row.get("prize_1", ""), bg)
        _data_style(ws1, r_idx, 5, row.get("prize_2", ""), bg)
        _data_style(ws1, r_idx, 6, row.get("prize_3", ""), bg)
        _data_style(ws1, r_idx, 7, row.get("prize_4", ""), bg)
        _data_style(ws1, r_idx, 8, row.get("prize_5", ""), bg)
        _data_style(ws1, r_idx, 9, row.get("prize_6", ""), bg)
        _data_style(ws1, r_idx, 10, row.get("prize_7", ""), bg)
        _data_style(ws1, r_idx, 11, row.get("prize_8", ""), bg)

    # ── Sheet 2: Tần suất số ────────────────────────────────────────────
    ws2 = wb.create_sheet("Tần Suất Số")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 18

    ws2.merge_cells("A1:C1")
    c = ws2["A1"]
    c.value = "THỐNG KÊ TẦN SUẤT SỐ LÔ TÔ"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2D6A4F")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    _header_style(ws2, 2, 1, "Số", "2D6A4F")
    _header_style(ws2, 2, 2, "Lần xuất hiện", "2D6A4F")
    _header_style(ws2, 2, 3, "Xếp hạng", "2D6A4F")

    sorted_freq = sorted(freq_data.items(), key=lambda x: x[1], reverse=True)
    max_count = sorted_freq[0][1] if sorted_freq else 1

    for i, (num, count) in enumerate(sorted_freq, 3):
        ratio = count / max_count
        bg = "F0FFF4" if ratio > 0.7 else ("FFFDE7" if ratio > 0.4 else "FFF5F5")
        rank = i - 2
        _data_style(ws2, i, 1, num, bg)
        _data_style(ws2, i, 2, count, bg)
        medal = "🥇" if rank == 1 else ("🥈" if rank == 2 else ("🥉" if rank == 3 else f"#{rank}"))
        _data_style(ws2, i, 3, medal, bg)

    # ── Sheet 3: Lô Gan ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Lô Gan")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 18

    ws3.merge_cells("A1:C1")
    c = ws3["A1"]
    c.value = "LÔ GAN — SỐ LÂU KHÔNG XUẤT HIỆN"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="C0392B")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    _header_style(ws3, 2, 1, "Số", "C0392B")
    _header_style(ws3, 2, 2, "Lần cuối xuất hiện", "C0392B")
    _header_style(ws3, 2, 3, "Số ngày vắng", "C0392B")

    for i, item in enumerate(gan_data, 3):
        days = item.get("days_absent", 0)
        bg = "FFF0F0" if days > 20 else ("FFFBE0" if days > 10 else "F5FFF5")
        _data_style(ws3, i, 1, item.get("number", ""), bg)
        _data_style(ws3, i, 2, item.get("last_date", ""), bg)
        _data_style(ws3, i, 3, days if days < 9999 else "Chưa xuất hiện", bg)

    # ── Sheet 4: Đầu Đuôi ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Đầu Đuôi")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:C1")
    c = ws4["A1"]
    c.value = "THỐNG KÊ ĐẦU ĐUÔI"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="7B2D8B")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 28

    ws4.column_dimensions["A"].width = 10
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 4
    ws4.column_dimensions["D"].width = 10
    ws4.column_dimensions["E"].width = 14

    _header_style(ws4, 2, 1, "Đầu số", "7B2D8B")
    _header_style(ws4, 2, 2, "Số lần", "7B2D8B")
    ws4["C2"] = ""
    _header_style(ws4, 2, 4, "Đuôi số", "7B2D8B")
    _header_style(ws4, 2, 5, "Số lần", "7B2D8B")

    head = head_tail.get("head", {})
    tail = head_tail.get("tail", {})
    for i in range(10):
        r = i + 3
        bg = "F8F0FF" if i % 2 == 0 else "FFFFFF"
        _data_style(ws4, r, 1, f"Đầu {i}", bg)
        _data_style(ws4, r, 2, head.get(str(i), 0), bg)
        _data_style(ws4, r, 4, f"Đuôi {i}", bg)
        _data_style(ws4, r, 5, tail.get(str(i), 0), bg)

    wb.save(output_path)
    return output_path
